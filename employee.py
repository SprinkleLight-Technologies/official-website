from flask import Flask, render_template, request, redirect, flash
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)
app.secret_key = "your_secret_key"

# Path to the Excel sheet
excel_file = "attendance.xlsx"

# Load or create the Excel workbook
try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = load_workbook("template.xlsx")
    sheet = workbook.active
    sheet.append(["Employee Name", "Check In Time", "Check Out Time", "Duration"])
    workbook.save(excel_file)

@app.route('/')
def dashboard():
    return render_template("employee.html")

@app.route('/check-in', methods=['POST'])
def check_in():
    name = "@User"  # Replace with dynamic user data in a real application
    check_in_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([name, check_in_time, "", ""])
    workbook.save(excel_file)
    flash("Check-in successful!", "success")
    return redirect('/')

@app.route('/check-out', methods=['POST'])
def check_out():
    name = "@User"  # Replace with dynamic user data in a real application
    check_out_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Find the latest check-in record without a check-out time
    for row in reversed(list(sheet.iter_rows(min_row=2, values_only=True))):
        if row[0] == name and not row[2]:  # No check-out time yet
            check_in_time = datetime.strptime(row[1], "%Y-%m-%d %H:%M:%S")
            duration = datetime.now() - check_in_time
            duration_str = str(duration).split('.')[0]
            sheet.append([name, row[1], check_out_time, duration_str])
            workbook.save(excel_file)
            flash("Check-out successful!", "success")
            return redirect('/')
    flash("No check-in record found!", "error")
    return redirect('/')

if __name__ == '__main__':
    app.run(debug=True)